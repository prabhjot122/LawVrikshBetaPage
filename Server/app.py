from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
from werkzeug.exceptions import BadRequest
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# CORS configuration
cors_origins = [
    "http://localhost:3000",
    "http://localhost:5173",
    "http://localhost:5174",
    "http://localhost:5175",
    "https://lawvrikshbetapage.onrender.com"
]

# Add production origins from environment variable
if os.environ.get('CORS_ORIGINS'):
    production_origins = os.environ.get('CORS_ORIGINS').split(',')
    cors_origins.extend([origin.strip() for origin in production_origins])

CORS(app, origins=cors_origins)

# Database configuration
DATABASE_URL = os.environ.get('DATABASE_URL')

if DATABASE_URL:
    # Production database (MySQL from Aiven or other providers)
    # Handle Aiven's MySQL URL format - convert mysql:// to mysql+pymysql://
    if DATABASE_URL.startswith('mysql://'):
        DATABASE_URL = DATABASE_URL.replace('mysql://', 'mysql+pymysql://', 1)

    # Handle SSL mode parameter for Aiven
    if '?ssl-mode=REQUIRED' in DATABASE_URL:
        DATABASE_URL = DATABASE_URL.replace('?ssl-mode=REQUIRED', '')
        # SSL will be handled by engine options below

    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
    logger.info('Using production MySQL database from DATABASE_URL')
else:
    # Local development - use MySQL with environment variables
    DB_HOST = os.environ.get('DB_HOST', 'localhost')
    DB_USER = os.environ.get('DB_USER', 'root')
    DB_PASSWORD = os.environ.get('DB_PASSWORD', '')
    DB_NAME = os.environ.get('DB_NAME', 'lawvriksh_db')
    DB_PORT = os.environ.get('DB_PORT', '3306')

    # Construct MySQL connection string with PyMySQL driver
    mysql_url = f'mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}'
    app.config['SQLALCHEMY_DATABASE_URI'] = mysql_url
    logger.info(f'Using local MySQL database: {DB_HOST}:{DB_PORT}/{DB_NAME}')

# MySQL-specific configuration
engine_options = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'pool_timeout': 20,
    'pool_size': 5,
    'max_overflow': 10,
}

# Add SSL configuration for production (Aiven)
if DATABASE_URL and ('aiven' in DATABASE_URL or 'ssl-mode' in os.environ.get('DATABASE_URL', '')):
    engine_options['connect_args'] = {'ssl': {}}

app.config['SQLALCHEMY_ENGINE_OPTIONS'] = engine_options

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Database Models
class UserRegistration(db.Model):
    __tablename__ = 'user_registrations'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False, index=True)
    email = db.Column(db.String(255), nullable=False, index=True)
    phone = db.Column(db.String(20), nullable=False)
    gender = db.Column(db.String(50), nullable=True)
    profession = db.Column(db.String(255), nullable=True)
    user_type = db.Column(db.String(20), nullable=False, index=True)  # 'USER' or 'Creator'

    # Metadata
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    ip_address = db.Column(db.String(45), nullable=True)  # Support IPv4 and IPv6
    user_agent = db.Column(db.Text, nullable=True)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'email': self.email,
            'phone': self.phone,
            'gender': self.gender,
            'profession': self.profession,
            'user_type': self.user_type,
            'submitted_at': self.submitted_at.isoformat() if self.submitted_at else None,
            'ip_address': self.ip_address,
            'user_agent': self.user_agent
        }
class Feedback(db.Model):
    __tablename__ = 'feedback'
    
    id = db.Column(db.Integer, primary_key=True)
    
    # Rating questions (1-5 scale)
    visual_design = db.Column(db.Integer, nullable=True)
    ease_of_navigation = db.Column(db.Integer, nullable=True)
    mobile_responsiveness = db.Column(db.Integer, nullable=True)
    overall_satisfaction = db.Column(db.Integer, nullable=True)
    ease_of_tasks = db.Column(db.Integer, nullable=True)
    quality_of_services = db.Column(db.Integer, nullable=True)
    
    # Conditional fields for low ratings
    visual_design_issue = db.Column(db.Text, nullable=True)
    ease_of_navigation_issue = db.Column(db.Text, nullable=True)
    mobile_responsiveness_issue = db.Column(db.Text, nullable=True)
    overall_satisfaction_issue = db.Column(db.Text, nullable=True)
    ease_of_tasks_issue = db.Column(db.Text, nullable=True)
    quality_of_services_issue = db.Column(db.Text, nullable=True)
    
    # Text area questions
    like_most = db.Column(db.Text, nullable=True)
    improvements = db.Column(db.Text, nullable=True)
    features = db.Column(db.Text, nullable=True)
    legal_challenges = db.Column(db.Text, nullable=True)
    additional_comments = db.Column(db.Text, nullable=True)
    
    # Follow-up questions
    contact_willing = db.Column(db.String(10), nullable=True)  # 'yes' or 'no'
    contact_email = db.Column(db.String(255), nullable=True)
    
    # Metadata
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    ip_address = db.Column(db.String(45), nullable=True)  # Support IPv6
    user_agent = db.Column(db.Text, nullable=True)
    
    def to_dict(self):
        return {
            'id': self.id,
            'visual_design': self.visual_design,
            'ease_of_navigation': self.ease_of_navigation,
            'mobile_responsiveness': self.mobile_responsiveness,
            'overall_satisfaction': self.overall_satisfaction,
            'ease_of_tasks': self.ease_of_tasks,
            'quality_of_services': self.quality_of_services,
            'visual_design_issue': self.visual_design_issue,
            'ease_of_navigation_issue': self.ease_of_navigation_issue,
            'mobile_responsiveness_issue': self.mobile_responsiveness_issue,
            'overall_satisfaction_issue': self.overall_satisfaction_issue,
            'ease_of_tasks_issue': self.ease_of_tasks_issue,
            'quality_of_services_issue': self.quality_of_services_issue,
            'like_most': self.like_most,
            'improvements': self.improvements,
            'features': self.features,
            'legal_challenges': self.legal_challenges,
            'additional_comments': self.additional_comments,
            'contact_willing': self.contact_willing,
            'contact_email': self.contact_email,
            'submitted_at': self.submitted_at.isoformat() if self.submitted_at else None,
            'ip_address': self.ip_address,
            'user_agent': self.user_agent
        }

def validate_feedback_data(data):
    """Validate feedback form data"""
    errors = []
    
    # Validate rating fields (should be 1-5 if provided)
    rating_fields = [
        'visualDesign', 'easeOfNavigation', 'mobileResponsiveness',
        'overallSatisfaction', 'easeOfTasks', 'qualityOfServices'
    ]
    
    for field in rating_fields:
        if field in data and data[field]:
            try:
                rating = int(data[field])
                if rating < 1 or rating > 5:
                    errors.append(f'{field} must be between 1 and 5')
            except (ValueError, TypeError):
                errors.append(f'{field} must be a valid number')
    
    # Validate conditional fields for low ratings
    conditional_mapping = {
        'visualDesign': 'visualDesignIssue',
        'easeOfNavigation': 'easeOfNavigationIssue',
        'mobileResponsiveness': 'mobileResponsivenessIssue',
        'overallSatisfaction': 'overallSatisfactionIssue',
        'easeOfTasks': 'easeOfTasksIssue',
        'qualityOfServices': 'qualityOfServicesIssue'
    }
    
    for rating_field, issue_field in conditional_mapping.items():
        if rating_field in data and data[rating_field]:
            try:
                rating = int(data[rating_field])
                if rating < 3 and (issue_field not in data or not data[issue_field].strip()):
                    errors.append(f'Please explain what you didn\'t like for {rating_field} (rating below 3)')
            except (ValueError, TypeError):
                pass  # Already handled above
    
    # Validate email if contact willing is yes
    if data.get('contactWilling') == 'yes':
        email = data.get('contactEmail', '').strip()
        if not email:
            errors.append('Email is required when willing to be contacted')
        elif '@' not in email or '.' not in email:
            errors.append('Please provide a valid email address')
    
    return errors

def generate_excel_report():
    """Generate Excel file with user registrations and feedback data"""
    try:
        # Create a new workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create User Registrations sheet
        ws1 = wb.create_sheet("User Registrations")

        # Headers for User Registrations
        headers1 = ['ID', 'Name', 'Email', 'Phone', 'Gender', 'Profession', 'User Type', 'Submitted At', 'IP Address']
        ws1.append(headers1)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Get user registration data
        registrations = UserRegistration.query.order_by(UserRegistration.submitted_at.desc()).all()
        for reg in registrations:
            ws1.append([
                reg.id,
                reg.name,
                reg.email,
                reg.phone,
                reg.gender or '',
                reg.profession or '',
                reg.user_type,
                reg.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if reg.submitted_at else '',
                reg.ip_address or ''
            ])

        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width

        # Create Feedback Submissions sheet
        ws2 = wb.create_sheet("Feedback Submissions")

        # Headers for Feedback
        headers2 = [
            'ID', 'Visual Design', 'Visual Design Issue', 'Ease of Navigation', 'Navigation Issue',
            'Mobile Responsiveness', 'Mobile Issue', 'Overall Satisfaction', 'Satisfaction Issue',
            'Ease of Tasks', 'Tasks Issue', 'Quality of Services', 'Services Issue',
            'Like Most', 'Improvements', 'Features', 'Legal Challenges', 'Additional Comments',
            'Contact Willing', 'Contact Email', 'Submitted At', 'IP Address'
        ]
        ws2.append(headers2)

        # Style headers
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Get feedback data
        feedback_list = Feedback.query.order_by(Feedback.submitted_at.desc()).all()
        for feedback in feedback_list:
            ws2.append([
                feedback.id,
                feedback.visual_design or '',
                feedback.visual_design_issue or '',
                feedback.ease_of_navigation or '',
                feedback.ease_of_navigation_issue or '',
                feedback.mobile_responsiveness or '',
                feedback.mobile_responsiveness_issue or '',
                feedback.overall_satisfaction or '',
                feedback.overall_satisfaction_issue or '',
                feedback.ease_of_tasks or '',
                feedback.ease_of_tasks_issue or '',
                feedback.quality_of_services or '',
                feedback.quality_of_services_issue or '',
                feedback.like_most or '',
                feedback.improvements or '',
                feedback.features or '',
                feedback.legal_challenges or '',
                feedback.additional_comments or '',
                feedback.contact_willing or '',
                feedback.contact_email or '',
                feedback.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if feedback.submitted_at else '',
                feedback.ip_address or ''
            ])

        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    except Exception as e:
        logger.error(f'Error generating Excel report: {str(e)}')
        return None

# API Routes
@app.route('/')
def home():
    """Root endpoint"""
    return jsonify({
        'message': 'LawVriksh Feedback API',
        'version': '1.0.0',
        'endpoints': {
            'health': '/api/health',
            'register': '/api/register',
            'feedback': '/api/feedback',
            'admin': '/admin'
        }
    })

@app.route('/admin')
def admin_dashboard():
    """Admin dashboard for managing data and downloading Excel files"""
    return render_template('admin.html')

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()})

@app.route('/api/register', methods=['POST'])
def register_user():
    """Register a new user (USER or Creator)"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Validate required fields
        required_fields = ['name', 'email', 'phone', 'userType']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400

        # Validate email format
        email = data.get('email', '').strip()
        if '@' not in email or '.' not in email:
            return jsonify({'error': 'Please provide a valid email address'}), 400

        # Validate user type
        user_type = data.get('userType')
        if user_type not in ['USER', 'Creator']:
            return jsonify({'error': 'User type must be USER or Creator'}), 400

        # Create user registration record
        registration = UserRegistration(
            name=data.get('name', '').strip(),
            email=email,
            phone=data.get('phone', '').strip(),
            gender=data.get('gender', '').strip() or None,
            profession=data.get('profession', '').strip() or None,
            user_type=user_type,
            ip_address=request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr),
            user_agent=request.headers.get('User-Agent')
        )

        db.session.add(registration)
        db.session.commit()

        # Generate updated Excel file (only save locally in development)
        if os.environ.get('FLASK_ENV') == 'development':
            excel_buffer = generate_excel_report()
            if excel_buffer:
                # Save Excel file to disk with error handling (development only)
                try:
                    with open('lawvriksh_data.xlsx', 'wb') as f:
                        f.write(excel_buffer.getvalue())
                except PermissionError:
                    logger.warning('Could not update Excel file - file may be open in another program')
                except Exception as e:
                    logger.error(f'Error saving Excel file: {str(e)}')

        logger.info(f'User registration submitted successfully with ID: {registration.id}')

        return jsonify({
            'message': 'Registration submitted successfully',
            'id': registration.id,
            'submitted_at': registration.submitted_at.isoformat()
        }), 201

    except Exception as e:
        db.session.rollback()
        logger.error(f'Error submitting registration: {str(e)}')
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/feedback', methods=['POST'])
def submit_feedback():
    """Submit feedback form"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate data
        validation_errors = validate_feedback_data(data)
        if validation_errors:
            return jsonify({'error': 'Validation failed', 'details': validation_errors}), 400
        
        # Create feedback record
        feedback = Feedback(
            visual_design=int(data.get('visualDesign')) if data.get('visualDesign') else None,
            ease_of_navigation=int(data.get('easeOfNavigation')) if data.get('easeOfNavigation') else None,
            mobile_responsiveness=int(data.get('mobileResponsiveness')) if data.get('mobileResponsiveness') else None,
            overall_satisfaction=int(data.get('overallSatisfaction')) if data.get('overallSatisfaction') else None,
            ease_of_tasks=int(data.get('easeOfTasks')) if data.get('easeOfTasks') else None,
            quality_of_services=int(data.get('qualityOfServices')) if data.get('qualityOfServices') else None,
            
            visual_design_issue=data.get('visualDesignIssue', '').strip() or None,
            ease_of_navigation_issue=data.get('easeOfNavigationIssue', '').strip() or None,
            mobile_responsiveness_issue=data.get('mobileResponsivenessIssue', '').strip() or None,
            overall_satisfaction_issue=data.get('overallSatisfactionIssue', '').strip() or None,
            ease_of_tasks_issue=data.get('easeOfTasksIssue', '').strip() or None,
            quality_of_services_issue=data.get('qualityOfServicesIssue', '').strip() or None,
            
            like_most=data.get('likeMost', '').strip() or None,
            improvements=data.get('improvements', '').strip() or None,
            features=data.get('features', '').strip() or None,
            legal_challenges=data.get('legalChallenges', '').strip() or None,
            additional_comments=data.get('additionalComments', '').strip() or None,
            
            contact_willing=data.get('contactWilling', '').strip() or None,
            contact_email=data.get('contactEmail', '').strip() or None,
            
            ip_address=request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr),
            user_agent=request.headers.get('User-Agent')
        )
        
        db.session.add(feedback)
        db.session.commit()

        # Generate updated Excel file (only save locally in development)
        if os.environ.get('FLASK_ENV') == 'development':
            excel_buffer = generate_excel_report()
            if excel_buffer:
                # Save Excel file to disk with error handling (development only)
                try:
                    with open('lawvriksh_data.xlsx', 'wb') as f:
                        f.write(excel_buffer.getvalue())
                except PermissionError:
                    logger.warning('Could not update Excel file - file may be open in another program')
                except Exception as e:
                    logger.error(f'Error saving Excel file: {str(e)}')

        logger.info(f'Feedback submitted successfully with ID: {feedback.id}')

        return jsonify({
            'message': 'Feedback submitted successfully',
            'id': feedback.id,
            'submitted_at': feedback.submitted_at.isoformat()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'Error submitting feedback: {str(e)}')
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/feedback', methods=['GET'])
def get_feedback():
    """Get all feedback (admin only - add authentication in production)"""
    try:
        # In production, add proper authentication here
        # For now, we'll add a simple API key check
        api_key = request.headers.get('X-API-Key')
        if api_key != os.environ.get('ADMIN_API_KEY', 'admin-key-123'):
            return jsonify({'error': 'Unauthorized'}), 401
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        if per_page > 100:
            per_page = 100
        
        feedback_query = Feedback.query.order_by(Feedback.submitted_at.desc())
        feedback_paginated = feedback_query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        return jsonify({
            'feedback': [f.to_dict() for f in feedback_paginated.items],
            'total': feedback_paginated.total,
            'pages': feedback_paginated.pages,
            'current_page': page,
            'per_page': per_page
        })
        
    except Exception as e:
        logger.error(f'Error retrieving feedback: {str(e)}')
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    """Download Excel file with all data (admin only)"""
    try:
        # Check authentication
        api_key = request.headers.get('X-API-Key')
        if api_key != os.environ.get('ADMIN_API_KEY', 'admin-key-123'):
            return jsonify({'error': 'Unauthorized'}), 401

        # Generate Excel file
        excel_buffer = generate_excel_report()
        if not excel_buffer:
            return jsonify({'error': 'Failed to generate Excel file'}), 500

        # Return file as download
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f'lawvriksh_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f'Error downloading Excel file: {str(e)}')
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/registrations', methods=['GET'])
def get_registrations():
    """Get all user registrations (admin only)"""
    try:
        # Check authentication
        api_key = request.headers.get('X-API-Key')
        if api_key != os.environ.get('ADMIN_API_KEY', 'admin-key-123'):
            return jsonify({'error': 'Unauthorized'}), 401

        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)

        if per_page > 100:
            per_page = 100

        registrations_query = UserRegistration.query.order_by(UserRegistration.submitted_at.desc())
        registrations_paginated = registrations_query.paginate(
            page=page, per_page=per_page, error_out=False
        )

        return jsonify({
            'registrations': [r.to_dict() for r in registrations_paginated.items],
            'total': registrations_paginated.total,
            'pages': registrations_paginated.pages,
            'current_page': page,
            'per_page': per_page
        })

    except Exception as e:
        logger.error(f'Error retrieving registrations: {str(e)}')
        return jsonify({'error': 'Internal server error'}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return jsonify({'error': 'Internal server error'}), 500

# Initialize database
def create_tables():
    """Create database tables"""
    try:
        with app.app_context():
            # Create all tables
            db.create_all()
            logger.info('Database tables created successfully using MySQL')
            logger.info('Tables created: user_registrations, feedback')

    except Exception as e:
        logger.error(f'Error creating database tables: {str(e)}')
        # Re-raise the exception so the app doesn't start with a broken database
        raise

if __name__ == '__main__':
    create_tables()
    port = int(os.environ.get('PORT', 3000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)
